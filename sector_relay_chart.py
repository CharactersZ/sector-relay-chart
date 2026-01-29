#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
赛道接力图生成器
每日自动获取行业板块涨幅数据，生成横向接力的Excel图表
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import akshare as ak
import os
import json
from collections import defaultdict

class SectorRelayChart:
    def __init__(self, excel_path="赛道接力图.xlsx", top_n=30):
        """
        初始化赛道接力图生成器
        
        Parameters:
        -----------
        excel_path : str
            Excel文件保存路径
        top_n : int
            每日取前N个板块，默认30
        """
        self.excel_path = excel_path
        self.top_n = top_n
        self.sector_colors = {}  # 存储每个板块的固定颜色
        self.color_palette = [
            "FFE6E6",  # 浅红
            "E6F3FF",  # 浅蓝
            "E6FFE6",  # 浅绿
            "FFF0E6",  # 浅橙
            "F0E6FF",  # 浅紫
            "FFFFE6",  # 浅黄
            "E6FFFF",  # 浅青
            "FFE6F0",  # 浅粉
            "E6FFE6",  # 浅绿2
            "F5F5DC",  # 米色
            "FFDAB9",  # 桃色
            "DDA0DD",  # 梅色
            "98FB98",  # 淡绿
            "F0E68C",  # 卡其色
            "FFB6C1",  # 浅粉红
            "B0E0E6",  # 粉蓝
            "FFA07A",  # 浅鲑鱼色
            "87CEEB",  # 天蓝色
            "D8BFD8",  # 蓟色
            "F5DEB3",  # 小麦色
        ]
        self.color_index = 0
        
    def get_sector_data(self, date_str=None):
        """
        获取指定日期的行业板块涨幅数据
        
        Parameters:
        -----------
        date_str : str
            日期字符串，格式：YYYYMMDD，如果为None则获取最新数据
            
        Returns:
        --------
        pd.DataFrame
            包含板块名称和涨幅的数据框
        """
        try:
            # 获取行业板块实时行情数据
            print(f"正在获取板块数据...")
            df = ak.stock_board_industry_name_em()
            
            # 如果指定了日期，需要获取历史数据（这里先获取最新数据）
            # 注意：akshare的行业板块数据是实时数据，历史数据需要其他接口
            # 这里先实现获取最新数据的功能
            
            # 选择需要的列：板块名称和涨跌幅
            if '板块名称' in df.columns and '涨跌幅' in df.columns:
                result_df = df[['板块名称', '涨跌幅']].copy()
            elif 'name' in df.columns and 'changePercent' in df.columns:
                result_df = df[['name', 'changePercent']].copy()
                result_df.columns = ['板块名称', '涨跌幅']
            else:
                # 尝试其他可能的列名
                print("数据列名:", df.columns.tolist())
                result_df = df.iloc[:, [0, 1]].copy()
                result_df.columns = ['板块名称', '涨跌幅']
            
            # 验证数据有效性
            if result_df.empty:
                raise ValueError("获取的数据为空")
            
            # 检查涨跌幅列是否有效
            if '涨跌幅' not in result_df.columns:
                raise ValueError("数据中缺少'涨跌幅'列")
            
            # 按涨跌幅降序排序
            result_df = result_df.sort_values('涨跌幅', ascending=False)
            
            # 验证涨跌幅数据是否合理（应该在-20%到20%之间，超出范围可能是数据错误）
            invalid_data = result_df[(result_df['涨跌幅'] < -20) | (result_df['涨跌幅'] > 20)]
            if not invalid_data.empty:
                print(f"⚠️  警告：发现异常涨跌幅数据 {len(invalid_data)} 条，已过滤")
                result_df = result_df[(result_df['涨跌幅'] >= -20) & (result_df['涨跌幅'] <= 20)]
            
            # 取前N个
            result_df = result_df.head(self.top_n).reset_index(drop=True)
            
            if result_df.empty:
                raise ValueError("过滤后数据为空，无法继续")
            
            # 显示前3名板块信息，确认数据真实性
            print(f"✓ 成功获取 {len(result_df)} 个真实板块数据")
            print(f"   前3名：{', '.join(result_df.head(3)['板块名称'].tolist())}")
            return result_df
            
        except Exception as e:
            print(f"❌ 获取行业板块数据失败: {e}")
            print("尝试使用备用方法（概念板块数据）...")
            # 备用方法：使用股票概念板块数据
            try:
                df = ak.stock_board_concept_name_em()
                if 'name' in df.columns and 'changePercent' in df.columns:
                    result_df = df[['name', 'changePercent']].copy()
                    result_df.columns = ['板块名称', '涨跌幅']
                    result_df = result_df.sort_values('涨跌幅', ascending=False)
                    result_df = result_df.head(self.top_n).reset_index(drop=True)
                    print(f"✓ 使用概念板块数据，成功获取 {len(result_df)} 个板块")
                    return result_df
                else:
                    raise ValueError("概念板块数据格式不正确")
            except Exception as e2:
                print(f"❌ 备用方法也失败: {e2}")
                raise Exception(
                    "无法获取真实板块数据！\n"
                    "可能的原因：\n"
                    "1. 网络连接问题，请检查网络\n"
                    "2. 数据源暂时不可用，请稍后重试\n"
                    "3. akshare库需要更新：pip install --upgrade akshare\n\n"
                    "程序已停止，不会使用虚假数据。"
                )
    
    
    def get_sector_color(self, sector_name):
        """
        为板块分配固定颜色
        
        Parameters:
        -----------
        sector_name : str
            板块名称
            
        Returns:
        --------
        str
            颜色代码（十六进制，不含#）
        """
        if sector_name not in self.sector_colors:
            # 分配新颜色
            color = self.color_palette[self.color_index % len(self.color_palette)]
            self.sector_colors[sector_name] = color
            self.color_index += 1
        return self.sector_colors[sector_name]
    
    def load_existing_data(self):
        """
        加载已存在的Excel数据
        
        Returns:
        --------
        dict
            包含历史数据的字典，格式：{日期: [板块列表]}
        """
        if not os.path.exists(self.excel_path):
            return {}
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            # 读取表头（日期）
            dates = []
            for col in range(2, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    dates.append(str(cell.value))
            
            # 读取数据
            data = {}
            for col_idx, date in enumerate(dates, start=2):
                sectors = []
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value:
                        sectors.append(str(cell.value))
                if sectors:
                    data[date] = sectors
            
            # 加载颜色映射
            color_file = os.path.join(os.path.dirname(self.excel_path), "sector_colors.json")
            if os.path.exists(color_file):
                with open(color_file, "r", encoding="utf-8") as f:
                    self.sector_colors = json.load(f)
            
            wb.close()
            return data
            
        except Exception as e:
            print(f"加载已有数据失败: {e}")
            return {}
    
    def save_color_mapping(self):
        """保存颜色映射到JSON文件"""
        # 颜色映射文件保存在Excel文件同目录
        color_file = os.path.join(os.path.dirname(self.excel_path), "sector_colors.json")
        with open(color_file, "w", encoding="utf-8") as f:
            json.dump(self.sector_colors, f, ensure_ascii=False, indent=2)
    
    def update_excel(self, date_str=None, auto_overwrite=False):
        """
        更新Excel文件，添加新的一列数据
        
        Parameters:
        -----------
        date_str : str
            日期字符串，格式：YYYY-MM-DD 或 YYYYMMDD，如果为None则使用今天
        auto_overwrite : bool
            如果数据已存在，是否自动覆盖，默认False（交互式询问）
        """
        # 获取今天的日期
        if date_str is None:
            today = datetime.now()
        else:
            try:
                if len(date_str) == 8:  # YYYYMMDD
                    today = datetime.strptime(date_str, "%Y%m%d")
                else:  # YYYY-MM-DD
                    today = datetime.strptime(date_str, "%Y-%m-%d")
            except:
                today = datetime.now()
        
        date_display = today.strftime("%Y-%m-%d")
        date_key = today.strftime("%Y%m%d")
        
        # 获取今日板块数据
        sector_df = self.get_sector_data(date_key)
        today_sectors = sector_df['板块名称'].tolist()
        
        # 加载已有数据
        existing_data = self.load_existing_data()
        
        # 检查今天的数据是否已存在
        if date_display in existing_data:
            if auto_overwrite:
                print(f"日期 {date_display} 的数据已存在，自动覆盖...")
            else:
                print(f"日期 {date_display} 的数据已存在，是否覆盖？")
                response = input("输入 'y' 覆盖，其他键跳过: ").strip().lower()
                if response != 'y':
                    print("跳过更新")
                    return
        
        # 添加今天的数据
        existing_data[date_display] = today_sectors
        
        # 保存颜色映射（从已有数据中恢复）
        for date, sectors in existing_data.items():
            for sector in sectors:
                self.get_sector_color(sector)
        
        # 创建Excel
        self.create_excel(existing_data)
        
        # 保存颜色映射
        self.save_color_mapping()
        
        print(f"✓ 已更新 {date_display} 的数据到Excel")
    
    def create_excel(self, data_dict):
        """
        创建或更新Excel文件
        
        Parameters:
        -----------
        data_dict : dict
            数据字典，格式：{日期: [板块列表]}
        """
        # 创建新的工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "赛道接力图"
        
        # 获取所有日期，按时间排序
        dates = sorted(data_dict.keys(), key=lambda x: datetime.strptime(x, "%Y-%m-%d"))
        
        # 写入表头
        ws.cell(row=1, column=1, value="排名")
        for col_idx, date in enumerate(dates, start=2):
            cell = ws.cell(row=1, column=col_idx, value=date)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 获取最大行数（所有日期中板块数量最多的）
        max_rows = max(len(sectors) for sectors in data_dict.values()) if data_dict else self.top_n
        
        # 写入数据和设置颜色
        for row in range(1, max_rows + 1):
            # 写入排名
            rank_cell = ws.cell(row=row + 1, column=1, value=row)
            rank_cell.font = Font(bold=True)
            rank_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入每个日期的板块数据
            for col_idx, date in enumerate(dates, start=2):
                sectors = data_dict.get(date, [])
                if row <= len(sectors):
                    sector_name = sectors[row - 1]
                    cell = ws.cell(row=row + 1, column=col_idx, value=sector_name)
                    
                    # 设置颜色
                    color = self.get_sector_color(sector_name)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8  # 排名列
        for col_idx in range(2, len(dates) + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12
        
        # 设置行高
        for row in range(1, max_rows + 2):
            ws.row_dimensions[row].height = 20
        
        # 冻结首行和首列
        ws.freeze_panes = 'B2'
        
        # 保存文件
        wb.save(self.excel_path)
        print(f"✓ Excel文件已保存: {self.excel_path}")


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='赛道接力图生成器 - 自动生成行业板块轮动接力图')
    parser.add_argument('-d', '--date', type=str, help='指定日期 (格式: YYYY-MM-DD 或 YYYYMMDD)，默认为今天')
    parser.add_argument('-f', '--file', type=str, default='赛道接力图.xlsx', help='Excel文件路径，默认: 赛道接力图.xlsx')
    parser.add_argument('-n', '--top-n', type=int, default=30, help='每日取前N个板块，默认: 30')
    parser.add_argument('--view', action='store_true', help='仅查看已有数据，不更新')
    parser.add_argument('--auto-overwrite', action='store_true', help='如果数据已存在，自动覆盖（用于定时任务）')
    
    args = parser.parse_args()
    
    print("=" * 50)
    print("赛道接力图生成器")
    print("=" * 50)
    
    # 创建生成器实例
    chart = SectorRelayChart(excel_path=args.file, top_n=args.top_n)
    
    if args.view:
        # 仅查看模式
        existing_data = chart.load_existing_data()
        if existing_data:
            print(f"\n已有数据日期: {', '.join(sorted(existing_data.keys()))}")
            print(f"共 {len(existing_data)} 天的数据")
        else:
            print("\n暂无历史数据")
    else:
        # 更新模式
        date_str = args.date
        if date_str:
            print(f"\n准备更新 {date_str} 的数据...")
        else:
            today = datetime.now().strftime("%Y-%m-%d")
            print(f"\n准备更新 {today} 的数据...")
        
        # 更新Excel
        chart.update_excel(date_str, auto_overwrite=args.auto_overwrite)
        
        print("\n完成！")
        print(f"Excel文件位置: {os.path.abspath(chart.excel_path)}")


if __name__ == "__main__":
    main()
